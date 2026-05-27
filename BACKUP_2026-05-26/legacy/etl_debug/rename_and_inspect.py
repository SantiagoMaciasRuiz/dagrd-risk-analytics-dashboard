"""
Renombra la hoja 'Simulacros 2024-2025' a 'Tablas dinámicas' en el Excel fuente
y luego inspecciona el archivo de COMITES para listar hojas y primeros encabezados.

Se ejecuta desde el entorno del proyecto (.venv).
"""
from pathlib import Path
import sys

base = Path("data/source")
source_candidates = list(base.glob("Reporte de actividades equipo social 2026*.xlsx"))
if not source_candidates:
    print("ERROR: no se encontró el archivo fuente en data/source")
    sys.exit(2)
source_file = source_candidates[0]
print("Archivo fuente:", source_file)

try:
    from openpyxl import load_workbook
except Exception as e:
    print("ERROR: openpyxl no disponible en el entorno. Instala en .venv.", e)
    sys.exit(3)

wb = load_workbook(source_file)
old_name = None
for name in wb.sheetnames:
    if name.lower().startswith("simulacros") or name.lower().startswith("simulacro"):
        old_name = name
        break
if old_name is None:
    print("No se encontró hoja 'Simulacros...' para renombrar. Hojas disponibles:", wb.sheetnames)
else:
    if 'Tablas dinámicas' in wb.sheetnames:
        print("Ya existe 'Tablas dinámicas' en el libro; no se renombrará '" + old_name + "'.")
    else:
        print(f"Renombrando hoja '{old_name}' → 'Tablas dinámicas'")
        ws = wb[old_name]
        ws.title = 'Tablas dinámicas'
        wb.save(source_file)
        print("Guardado.")

# Inspeccionar archivo de COMITES
comites_path = Path(r"data/model/CONSOLIDADO COMITES COMISIONES 03-2026_Construcciom.xlsx")
if not comites_path.exists():
    print("ERROR: no se encontró el archivo de comités:", comites_path)
    sys.exit(0)

wb2 = load_workbook(comites_path, read_only=True)
print('\nArchivo comités:', comites_path)
print('Hojas:', wb2.sheetnames)

for s in wb2.sheetnames:
    ws = wb2[s]
    first_row = next(ws.iter_rows(values_only=True), None)
    print(f"\nHoja: {s}")
    if first_row:
        headers = [str(c) if c is not None else '' for c in first_row]
        print('Encabezados (primer fila):', headers[:20])
    else:
        print('Hoja vacía o sin fila 1')

print('\nInspección completada.')

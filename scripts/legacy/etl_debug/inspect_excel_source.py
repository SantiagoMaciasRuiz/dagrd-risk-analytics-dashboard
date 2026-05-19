"""
Inspecciona el archivo Excel fuente y muestra:
- lista de hojas
- número de filas (estimado) y columnas por hoja
- primeras filas y encabezados (hasta 10)
- búsqueda de hojas/columnas de interés: 'Tablas dinámicas', 'Simulacros', 'CAM', 'SAT-C', 'Sheet1'

Se ejecuta dentro del entorno del proyecto (.venv recomendado).
"""
import sys
from pathlib import Path
base = Path("data/source")
candidates = list(base.glob("Reporte de actividades equipo social 2026*.xlsx"))
if not candidates:
    print(f"ERROR: no se encontró ningún archivo que empiece con 'Reporte de actividades equipo social 2026' en {base}")
    sys.exit(2)
p = candidates[0]
print(f"Usando archivo: {p}")

try:
    import pandas as pd
except Exception:
    pd = None

sheets = []
if pd is not None:
    try:
        xls = pd.ExcelFile(p)
        sheets = xls.sheet_names
    except Exception:
        sheets = []

if not sheets:
    # fallback a openpyxl
    try:
        from openpyxl import load_workbook
        wb = load_workbook(p, read_only=True)
        sheets = wb.sheetnames
    except Exception as e:
        print("ERROR leyendo el Excel:", e)
        sys.exit(3)

print("Hojas encontradas:")
for s in sheets:
    print(f"- {s}")

keywords = ['Tablas dinámicas', 'Tablas dinamicas', 'Simulacros', 'CAM', 'SAT-C', 'Sheet1']
print('\nResumen por hoja:')
for s in sheets:
    print(f"\n--- Hoja: {s}")
    try:
        if pd is not None:
            df = pd.read_excel(p, sheet_name=s, nrows=5)
            # conteo aproximado: leer solo cabezera y contar no-nulos en primera columna para estimar filas
            try:
                full = pd.read_excel(p, sheet_name=s)
                filas = len(full)
            except Exception:
                filas = 'desconocido'
            cols = list(df.columns)
            sample = df.head(3).astype(object).where(pd.notnull(df), None).values.tolist()
        else:
            from openpyxl import load_workbook
            wb = load_workbook(p, read_only=True)
            ws = wb[s]
            rows = []
            cols = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i==0:
                    cols = [str(c) if c is not None else '' for c in row]
                if i<3:
                    rows.append(list(row))
                if i>1000:
                    break
            filas = 'desconocido'
            sample = rows
        print(f"Columnas (n={len(cols)}): {cols[:10]}")
        print(f"Filas (estimado): {filas}")
        print("Muestra primeras filas:")
        for r in sample:
            print("  ", r)
        # busquedas
        lower_cols = [str(c).lower() for c in cols]
        found = []
        for k in ['cam','simulacros','tablas','sat-c','sat c','sheet1']:
            if any(k in (c or '').lower() for c in cols):
                found.append(k)
        if found:
            print("Columnas de interés detectadas (fragmentos):", found)
    except Exception as e:
        print("  ERROR al leer la hoja:", e)

# búsqueda global: ¿existen exactamente las hojas buscadas?
print('\nPresencia de hojas clave:')
for k in ['Tablas dinámicas','Simulacros','CAM','SAT-C','Sheet1']:
    print(f"- {k}:", "Sí" if any(k.lower()==s.lower() for s in sheets) else "No (o renombrada)")

# búsqueda rápida de columnas CAM en todas las hojas
print('\nBúsqueda de columna con nombre que contenga "cam" en todas las hojas:')
for s in sheets:
    try:
        if pd is not None:
            df = pd.read_excel(p, sheet_name=s, nrows=0)
            cols = list(df.columns)
        else:
            from openpyxl import load_workbook
            wb = load_workbook(p, read_only=True)
            ws = wb[s]
            first = next(ws.iter_rows(values_only=True), [])
            cols = [str(c) if c is not None else '' for c in first]
        if any('cam' in (c or '').lower() for c in cols):
            print(f"- Hoja {s}: contiene columna con 'cam' -> {[c for c in cols if 'cam' in (c or '').lower()]}" )
    except Exception:
        pass

print('\nInspección completada.')

from openpyxl import load_workbook
import pandas as pd
p='data/model/Modelo_Reporte_Paginas_2026.xlsx'
print('Workbook:',p)
wb = load_workbook(p, read_only=True)
print('Sheets:', wb.sheetnames)

sheets_to_check = ['Dim_SATC','Dim_SATC_Relaciones','Dim_Semilleros','Dim_Semilleros_Confiable']
for s in sheets_to_check:
    try:
        df = pd.read_excel(p, sheet_name=s, engine='openpyxl')
        print('\n---', s, f'rows={len(df)} cols={len(df.columns)}')
        print(df.head(5).to_string(index=False))
        print('columns:', list(df.columns))
    except Exception as e:
        print('\n---', s, 'MISSING or ERROR:', e)

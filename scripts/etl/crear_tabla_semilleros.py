#!/usr/bin/env python3
"""
Crear tabla confiable de Semilleros DAGRD
Integra datos de Semilleros en el modelo Excel principal
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

# Datos de Semilleros proporcionados por DAGRD
SEMILLEROS_DATA = [
    {"N": 1, "Semillero": "Semillero IE Fe y Alegría", "Comuna": 1, "Comuna_Nombre": "Popular", "Barrio_Organizacion": "Popular - IE Fe y Alegría"},
    {"N": 2, "Semillero": "Semillero Comunidad Embera Dobida", "Comuna": 80, "Comuna_Nombre": "Corregimiento de San Antonio de Prado", "Barrio_Organizacion": "San Antonio de Prado"},
    {"N": 3, "Semillero": "Semillero Olaya Herrera", "Comuna": 7, "Comuna_Nombre": "Robledo", "Barrio_Organizacion": "Olaya Herrera - Comisión de Gestión del Riesgo Olaya Herrera"},
    {"N": 4, "Semillero": "Semillero Villatina", "Comuna": 8, "Comuna_Nombre": "Villa Hermosa", "Barrio_Organizacion": "Villatina"},
    {"N": 5, "Semillero": "Semillero El Pesebre", "Comuna": 13, "Comuna_Nombre": "San Javier", "Barrio_Organizacion": "El Pesebre - Mesa de gestión del Riesgo"},
    {"N": 6, "Semillero": "Semillero Centro Integrado San Cristóbal", "Comuna": 60, "Comuna_Nombre": "Corregimiento de San Cristóbal", "Barrio_Organizacion": "Centralidad  - Unidad de discapacidad, Centro Integrado"},
    {"N": 7, "Semillero": "Semillero IE Ciudadela Nuevo Occidente Pedregal Bajo", "Comuna": 60, "Comuna_Nombre": "Corregimiento de San Cristóbal", "Barrio_Organizacion": "Pedregal Bajo - IE Ciudadela Nuevo Occidente"},
    {"N": 8, "Semillero": "Semillero IE Ciudadela Nuevo Occidente", "Comuna": 60, "Comuna_Nombre": "Corregimiento de San Cristóbal", "Barrio_Organizacion": "Nuevo Occidente - IE Ciudadela Nuevo Occidente"},
    {"N": 9, "Semillero": "Semillero IE Manzanillo", "Comuna": 70, "Comuna_Nombre": "Corregimiento de Altavista", "Barrio_Organizacion": "Vereda San José de Manzanillo - IE Manzanillo"},
    {"N": 10, "Semillero": "Semillero La Isla", "Comuna": 2, "Comuna_Nombre": "Santa Cruz", "Barrio_Organizacion": "La Isla - Corporación Progreso Social"},
    {"N": 11, "Semillero": "Semillero Moravia Oasis", "Comuna": 4, "Comuna_Nombre": "Aranjuez", "Barrio_Organizacion": "Jac Moravia sector Oasis Tropical"},
    {"N": 12, "Semillero": "Semillero Moravia El Bosque", "Comuna": 4, "Comuna_Nombre": "Aranjuez", "Barrio_Organizacion": "Jac Moravia sector el Bosque"},
    {"N": 13, "Semillero": "Semillero Palermo", "Comuna": 4, "Comuna_Nombre": "Aranjuez", "Barrio_Organizacion": "Jac Palermo"},
    {"N": 14, "Semillero": "Semillero San Cayetano", "Comuna": 4, "Comuna_Nombre": "Aranjuez", "Barrio_Organizacion": "Jac San Cayetano"},
    {"N": 15, "Semillero": "Semillero Álamos", "Comuna": 4, "Comuna_Nombre": "Aranjuez", "Barrio_Organizacion": "Jac Álamos"},
    {"N": 16, "Semillero": "Semillero San Nicolas", "Comuna": 4, "Comuna_Nombre": "Aranjuez", "Barrio_Organizacion": "Jac San Nicolas"},
    {"N": 17, "Semillero": "Semillero Campo Valdés El Calvario", "Comuna": 4, "Comuna_Nombre": "Aranjuez", "Barrio_Organizacion": "Jac Campo Valdés el Calvario"},
    {"N": 18, "Semillero": "Semillero Sevilla", "Comuna": 4, "Comuna_Nombre": "Aranjuez", "Barrio_Organizacion": "Jac Sevilla"},
    {"N": 19, "Semillero": "Semillero Miranda", "Comuna": 4, "Comuna_Nombre": "Aranjuez", "Barrio_Organizacion": "Jac Miranda"},
    {"N": 20, "Semillero": "Semillero Emisora La Cuarta Estación", "Comuna": 4, "Comuna_Nombre": "Aranjuez", "Barrio_Organizacion": "San Pedro - Emisora la Cuarta Estación"},
]

def crear_tabla_semilleros():
    """Crear tabla Dim_Semilleros en Excel"""
    
    excel_file = Path("data/model/Modelo_Reporte_Paginas_2026.xlsx")
    
    if not excel_file.exists():
        print(f"❌ No se encontró {excel_file}")
        return
    
    # Cargar workbook
    wb = openpyxl.load_workbook(excel_file)
    
    # Eliminar hoja si existe
    if "Dim_Semilleros" in wb.sheetnames:
        del wb["Dim_Semilleros"]
    
    # Crear nueva hoja
    ws = wb.create_sheet("Dim_Semilleros", 1)  # Inserta como segunda hoja
    
    # Crear DataFrame
    df = pd.DataFrame(SEMILLEROS_DATA)
    
    # Escribir headers
    headers = ["N°", "Semillero", "Comuna", "Comuna_Nombre", "Barrio_Organización"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Escribir datos
    for row_idx, record in enumerate(SEMILLEROS_DATA, 2):
        ws.cell(row=row_idx, column=1, value=record["N"])
        ws.cell(row=row_idx, column=2, value=record["Semillero"])
        ws.cell(row=row_idx, column=3, value=record["Comuna"])
        ws.cell(row=row_idx, column=4, value=record["Comuna_Nombre"])
        ws.cell(row=row_idx, column=5, value=record["Barrio_Organizacion"])
    
    # Ajustar ancho de columnas
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 50
    
    # Aplicar bordes
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    for row in ws.iter_rows(min_row=1, max_row=len(SEMILLEROS_DATA)+1, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="center")
    
    # Congelar primera fila
    ws.freeze_panes = "A2"
    
    # Guardar
    wb.save(excel_file)
    print(f"✓ Tabla Dim_Semilleros creada exitosamente")
    print(f"  - {len(SEMILLEROS_DATA)} semilleros registrados")
    print(f"  - Distribuidos en {df['Comuna'].nunique()} comunas")
    print(f"  - Guardado en: {excel_file}")

if __name__ == "__main__":
    crear_tabla_semilleros()

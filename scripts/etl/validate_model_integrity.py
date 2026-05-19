#!/usr/bin/env python3
from openpyxl import load_workbook

try:
    wb = load_workbook('data/model/Modelo_Reporte_Paginas_2026.xlsx', read_only=True)
    print('✓ Archivo válido (sin corrupción)')
    print(f'Total hojas: {len(wb.sheetnames)}')
    print(f'\nHojas de hechos presentes:')
    print(f'  Hecho_Participacion_General: {"Hecho_Participacion_General" in wb.sheetnames}')
    print(f'  Hecho_Simulacros: {"Hecho_Simulacros" in wb.sheetnames}')
    print(f'  Hecho_Demografia: {"Hecho_Demografia" in wb.sheetnames}')
    print(f'  Dim_Comites_Comisiones_2026: {"Dim_Comites_Comisiones_2026" in wb.sheetnames}')
except Exception as e:
    print(f'✗ Error: {e}')

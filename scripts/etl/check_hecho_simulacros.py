#!/usr/bin/env python3
from openpyxl import load_workbook

wb = load_workbook('data/model/Modelo_Reporte_Paginas_2026.xlsx', read_only=True, data_only=True)

if 'Hecho_Simulacros' in wb.sheetnames:
    ws = wb['Hecho_Simulacros']
    # Leer encabezado de la primera fila
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
    
    print(f'✓ Hecho_Simulacros existe')
    print(f'Columnas ({len(headers)}):')
    for i, h in enumerate(headers, 1):
        print(f'  {i}. {h}')
    
    # Contar filas de datos
    data_rows = ws.max_row - 1  # excluyendo header
    print(f'\nFilas de datos: {data_rows}')
    
    # Verificar si tiene fila_origen
    if 'fila_origen' in headers:
        print(f'\n✓ Columna "fila_origen" EXISTE')
    else:
        print(f'\n✗ Columna "fila_origen" FALTA')
else:
    print('✗ Hecho_Simulacros NO EXISTE')

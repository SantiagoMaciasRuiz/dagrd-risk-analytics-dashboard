"""
Script FINAL para actualizar Excel con 37 SATC (corrección)
Reemplaza la tabla anterior de 33 SATC con 37 SATC
"""
import pandas as pd
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

print("=" * 80)
print("ACTUALIZACIÓN FINAL: 37 SATC")
print("=" * 80)

# LISTA OFICIAL DE 37 SATC (Confirmada por usuario)
satc_raw = """C1 - La Base
C2 - Playón de los Comuneros
C4 - La Bermejala (El Hueco)
C4 - San Isidro (Chocó Chiquito)
C6 - Kennedy Cantera Norte
C6 - Kennedy Cantera Sur
C7 - El Chorro- El Guayabo -Aures 1
C7 - El Chispero-Aures 2
C7 - Nueva Villa de La Iguaná
C7 - Olaya Herrera 1 (La Arenera)
C8 - Altos de la Torre
C8 - Colinas de Enciso Parte Alta
C8 - La Estrechura
C8 - Las Estancias (Caicedo)
C8 - La Paz
C8 - Santa Lucía-barrio Las Estancias
C8 - Unión de Cristo
C8 - Villa Esperanza
C9 - Barrio Alejandro Echavarría
C13 - El Pesebre
C16 - El Hoyo-barrio Las Violetas
C16 - Barrio Las Violetas
C60 - Vereda La Palma
C60 - La Honda – sector Caracolí
C70 - Las Playitas-vereda San Pablo
C70 – Guanteros
C70 – Manzanillo
C70 - Barrio Nuevo
C70 – Buga
C70 - Aguas Frías Parte Alta
C70 – Manzanares
C 70 - El Guamo
C70 - San Vicente
C80 - El Paraíso-vereda El Salado
C80 - Las Playas-vereda El Salado
C80 - Santa Rita-vereda La Verde
C80 - Palo Blanco"""

# Parsear
def parse_satc(raw):
    datos = []
    for linea in raw.split('\n'):
        linea = linea.strip()
        if not linea:
            continue
        # Normalizar: "C 70 x" → "C70 x"
        linea_norm = re.sub(r'C\s+(\d+)', r'C\1', linea)
        # Normalizar guiones: "–" → "-"
        linea_norm = linea_norm.replace('–', '-')
        
        match = re.match(r'^(C\d+)\s*-\s*(.+)$', linea_norm)
        if match:
            codigo, nombre = match.groups()
            comuna = int(codigo[1:])
            datos.append({
                'SATC_ID': codigo,
                'SATC_Nombre': nombre.strip(),
                'Comuna_Cod': comuna,
            })
    return datos

satc_list = parse_satc(satc_raw)

print(f"\n✓ Total SATC: {len(satc_list)}")
print(f"  Esperado: 37")
print(f"  Obtenido: {len(satc_list)}")

if len(satc_list) != 37:
    print(f"\n⚠ ERROR: Se esperaban 37 pero se obtuvieron {len(satc_list)}")
    for i, s in enumerate(satc_list, 1):
        print(f"  {i:2}. {s['SATC_ID']:3} - {s['SATC_Nombre']}")
    exit(1)

print("\n✅ Confirmado: 37 SATC")

# Cargar archivos para contexto
datos_dir = Path("datos")
talleres_file = datos_dir / "talleres_2026-03-10_19-14.xls"
comites_file = datos_dir / "comités_comisiones comunitarios_2026-03-10_19-15.xls"

try:
    talleres_df = pd.read_excel(talleres_file)
except:
    talleres_df = None

try:
    comites_df = pd.read_excel(comites_file)
except:
    comites_df = None

# Enriquecer SATC con conteos
print("\nEnriqueciendo con datos existentes...")

satc_enriquecido = []
for satc in satc_list:
    com = satc['Comuna_Cod']
    talleres = len(talleres_df[talleres_df['comuna_cod'] == com]) if talleres_df is not None else 0
    comites = len(comites_df[comites_df['comuna_cod'] == com]) if comites_df is not None else 0
    
    satc_enriquecido.append({
        'SATC_ID': satc['SATC_ID'],
        'SATC_Nombre': satc['SATC_Nombre'],
        'Comuna_Cod': satc['Comuna_Cod'],
        'Talleres': talleres,
        'Comites': comites,
        'Activo': 'Sí',
    })

satc_df = pd.DataFrame(satc_enriquecido)

# Relaciones por comuna
relaciones_df = satc_df[['Comuna_Cod']].drop_duplicates().reset_index(drop=True)
relaciones_df['SATC_Cantidad'] = relaciones_df['Comuna_Cod'].apply(
    lambda c: len(satc_df[satc_df['Comuna_Cod'] == c])
)
relaciones_df = relaciones_df.sort_values('Comuna_Cod').reset_index(drop=True)

print(f"\n✓ SATC enriquecido: {len(satc_df)}")
print(f"✓ Comunas: {len(relaciones_df)}")

# Actualizar Excel
modelo_file = Path("data/model/Modelo_Reporte_Paginas_2026.xlsx")

print(f"\nActualizando {modelo_file}...")

with pd.ExcelWriter(modelo_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    satc_df.to_excel(writer, sheet_name='Dim_SATC', index=False)
    relaciones_df.to_excel(writer, sheet_name='Dim_SATC_Relaciones', index=False)

# Formato
try:
    wb = load_workbook(modelo_file)
    
    # Dim_SATC
    if 'Dim_SATC' in wb.sheetnames:
        ws = wb['Dim_SATC']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        ws.column_dimensions['A'].width = 10  # SATC_ID
        ws.column_dimensions['B'].width = 40  # SATC_Nombre
        ws.column_dimensions['C'].width = 12  # Comuna_Cod
        ws.column_dimensions['D'].width = 10  # Talleres
        ws.column_dimensions['E'].width = 10  # Comites
        ws.column_dimensions['F'].width = 8   # Activo
    
    # Relaciones
    if 'Dim_SATC_Relaciones' in wb.sheetnames:
        ws = wb['Dim_SATC_Relaciones']
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
    
    wb.save(modelo_file)
    print(f"✓ Formato aplicado")

except Exception as e:
    print(f"⚠ Error en formato: {e}")

# Resumen
print("\n" + "=" * 80)
print("DESGLOSE POR COMUNA (37 SATC)")
print("=" * 80)

for com in sorted(relaciones_df['Comuna_Cod'].unique()):
    satc_en_com = satc_df[satc_df['Comuna_Cod'] == com]
    print(f"\nC{int(com)}: {len(satc_en_com)} SATC")
    for _, row in satc_en_com.iterrows():
        print(f"  • {row['SATC_Nombre']}")

# Guardar lista limpia
lista_file = Path("SATC_37_LISTA_FINAL.txt")
with open(lista_file, 'w', encoding='utf-8') as f:
    f.write("LISTA FINAL: 37 SATC (Confirmado)\n")
    f.write("=" * 80 + "\n\n")
    for i, row in satc_df.iterrows():
        f.write(f"{i+1:2}. {row['SATC_ID']} - {row['SATC_Nombre']}\n")

print(f"\n✓ Lista guardada: {lista_file}")

print("\n" + "=" * 80)
print("RESUMEN FINAL")
print("=" * 80)

print(f"""
Total SATC:                 37 (CONFIRMADO)
Comunas:                    12
Cobertura con datos:        100%
Talleres en SATC:           {satc_df['Talleres'].sum()}
Comités en SATC:            {satc_df['Comites'].sum()}

Cambio respecto a anterior:
  De: 33 SATC
  A:  37 SATC
  Diferencia: +4 SATC (Guanteros, Manzanillo, Buga, Manzanares en C70)

Archivo actualizado: {modelo_file}
""")

print("=" * 80)

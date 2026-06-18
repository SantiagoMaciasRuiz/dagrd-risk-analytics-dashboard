from openpyxl import load_workbook
from pathlib import Path


def _resolve_source_file() -> Path:
    source_dir = Path("data/source")
    candidates = sorted(
        source_dir.glob("Reporte de actividades equipo social 2026*.xlsx"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if candidates:
        return candidates[0]
    return source_dir / "Reporte de actividades equipo social 2026.xlsx"


source_file = _resolve_source_file()

# Cargar archivo
wb = load_workbook(source_file)
ws = wb['Sheet1']

print(f"Hoja cargada: {ws.title}")
print(f"Dimensiones: {ws.dimensions}")

cambios = 0

# Hacer cambios en celdas específicas
# Buscar la columna "Indique el número de personas participantes"
header_row = 1
header_col = None

for col_idx, cell in enumerate(ws[header_row], start=1):
    if cell.value and "Indique el número de personas participantes" in str(cell.value):
        header_col = col_idx
        print(f"Columna encontrada: {cell.value} en columna {col_idx}")
        # Cambiar algunos valores en esta columna
        for row_idx in range(2, min(7, ws.max_row)):
            current_val = ws.cell(row=row_idx, column=col_idx).value
            if current_val:
                try:
                    new_val = int(float(current_val)) + 5
                    ws.cell(row=row_idx, column=col_idx).value = new_val
                    cambios += 1
                    print(f"  Fila {row_idx}: {current_val} → {new_val}")
                except:
                    pass
        break

# Buscar columna "Mujeres"
header_col = None
for col_idx, cell in enumerate(ws[header_row], start=1):
    if cell.value and str(cell.value).strip() == "Mujeres":
        header_col = col_idx
        print(f"Columna encontrada: {cell.value} en columna {col_idx}")
        # Cambiar algunos valores
        for row_idx in range(2, min(7, ws.max_row)):
            current_val = ws.cell(row=row_idx, column=col_idx).value
            if current_val:
                try:
                    new_val = int(float(current_val)) + 3
                    ws.cell(row=row_idx, column=col_idx).value = new_val
                    cambios += 1
                    print(f"  Fila {row_idx}: {current_val} → {new_val}")
                except:
                    pass
        break

# Guardar archivo
wb.save(source_file)
print(f"\n✅ Archivo actualizado: {source_file}")
print(f"   Total de cambios realizados: {cambios}")

#!/usr/bin/env python3
from openpyxl import load_workbook

wb = load_workbook('data/model/Modelo_Reporte_Paginas_2026.xlsx', read_only=True)

for sheet_name in ['Hecho_Participacion_General', 'Hecho_Simulacros', 'Dim_SATC_Relaciones']:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        print(f"\n{sheet_name}:")
        print(f"  Columnas: {headers[:10]}...")  # primeras 10
        print(f"  Total: {len([h for h in headers if h])}")
    else:
        print(f"\n{sheet_name}: NO EXISTE")

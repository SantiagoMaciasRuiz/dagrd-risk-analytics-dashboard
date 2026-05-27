#!/usr/bin/env python3
from __future__ import annotations

import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

BASE = Path(r"c:\Users\santi\OneDrive\Escritorio\Chamba\Dashboard")
MODELO = BASE / "data" / "model" / "Modelo_Reporte_Paginas_2026.xlsx"
SATC_VALID = BASE / "data" / "model" / "validacion_satc_2026_03_18.xlsx"
ENTREGABLE_MODELO = BASE / "entregables" / "publicacion_nube_2026-03-20" / "02_datos_modelo" / "Modelo_Reporte_Paginas_2026.xlsx"
DIM_COMITES_CSV = BASE / "data" / "model" / "Dim_Comites_Comisiones_2026.csv"
EXCEL_MAESTRO_POWERBI = BASE / "data" / "model" / "Excel_Maestro_PowerBI.xlsx"
CONSOLIDADO_COMITES_PREFERIDO = BASE / "data" / "model" / "CONSOLIDADO COMITES COMISIONES 03-2026.xlsx"
CONSOLIDADO_COMITES_LEGACY = BASE / "data" / "model" / "CONSOLIDADO COMITES COMISIONES 03-2026_Construcciom.xlsx"


def _get_consolidado_comites_path() -> Path:
    # Prioriza el archivo solicitado por el usuario; mantiene compatibilidad con nombre legacy.
    if CONSOLIDADO_COMITES_PREFERIDO.exists():
        return CONSOLIDADO_COMITES_PREFERIDO
    return CONSOLIDADO_COMITES_LEGACY


def _strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value)
    return "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")


def _normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def _normalize_comite_name(value: object) -> str:
    text = _normalize_spaces(str(value or "").replace("\xa0", " "))
    if not text or text.lower() in {"nan", "none"}:
        return ""
    text = _strip_accents(text).upper()
    text = re.sub(r"[;|,]+$", "", text).strip()
    return text

def _normalize_header(value: object) -> str:
    text = _normalize_spaces(str(value or "").replace("\xa0", " ").replace("\n", " "))
    text = _strip_accents(text).lower()
    return text


def _extract_code_from_sheet_name(sheet_name: str) -> int | None:
    match = re.search(r"(\d{1,3})", str(sheet_name or ""))
    if match:
        return int(match.group(1))
    return None


def _parse_comuna_code(raw_code: object, raw_name: object) -> int | None:
    code = pd.to_numeric(raw_code, errors="coerce")
    if pd.notna(code) and int(code) > 0:
        return int(code)

    name = _strip_accents(_normalize_spaces(str(raw_name or "").lower()))
    if not name or name in {"nan", "none"}:
        return None

    match = re.search(r"\b(\d{1,3})\b", name)
    if match:
        return int(match.group(1))

    correg_map = {
        "palmitas": 50,
        "san cristobal": 60,
        "altavista": 70,
        "san antonio de prado": 80,
        "santa elena": 90,
    }
    for key, value in correg_map.items():
        if key in name:
            return value

    comuna_map = {
        "popular": 1,
        "santa cruz": 2,
        "manrique": 3,
        "aranjuez": 4,
        "castilla": 5,
        "doce de octubre": 6,
        "robledo": 7,
        "villa hermosa": 8,
        "buenos aires": 9,
        "la candelaria": 10,
        "laureles estadio": 11,
        "la america": 12,
        "san javier": 13,
        "el poblado": 14,
        "guayabal": 15,
        "belen": 16,
    }
    for key, value in comuna_map.items():
        if key in name:
            return value
    return None


def _build_dim_comites_desde_consolidado_confiable() -> pd.DataFrame | None:
    expected_cols = ["comuna_cod", "comite_comision_nombre", "comite_comision_etiqueta"]
    consolidado_path = _get_consolidado_comites_path()
    if not consolidado_path.exists():
        return None

    try:
        xls = pd.ExcelFile(consolidado_path)
    except Exception:
        return None

    rows: list[dict[str, object]] = []
    for sheet in xls.sheet_names:
        sh = _normalize_header(sheet)
        if "resumen" in sh or "coordinadores" in sh or sh == "consolidado":
            continue

        try:
            df_raw = pd.read_excel(consolidado_path, sheet_name=sheet, header=None)
        except Exception:
            continue

        # Buscar el header row (contiene "Comuna2" y "NombreGCRD" o similar)
        header_idx = None
        header_map = {}
        for i in range(min(8, len(df_raw))):
            vals = [str(v).strip() for v in df_raw.iloc[i].tolist()]
            norm = [v.lower().replace(" ", "") for v in vals]
            if "comuna2" in norm and any("nombreccgrd" in x for x in norm):
                header_idx = i
                header_map = {j: vals[j] for j in range(len(vals))}
                break

        if header_idx is None:
            continue

        # Leer datos desde la fila después del header
        data = df_raw.iloc[header_idx + 1:].copy()
        data.columns = [header_map.get(j, f"col_{j}") for j in range(data.shape[1])]
        cols_norm = {str(c).strip().lower().replace(" ", ""): c for c in data.columns}

        # Encontrar columnas clave
        col_tipo = cols_norm.get("comuna")  # Columna que contiene T/S
        col_comuna2 = cols_norm.get("comuna2")  # Código de comuna
        col_nombre = None
        for k, c in cols_norm.items():
            if "nombreccgrd" in k:
                col_nombre = c
                break

        if not (col_tipo and col_comuna2 and col_nombre):
            continue

        # Procesar datos
        work = data[[col_tipo, col_comuna2, col_nombre]].copy()
        work.columns = ["tipo_raw", "comuna2", "nombre_ccgrd"]
        
        # Normalizar tipo
        work["tipo_raw"] = work["tipo_raw"].astype(str).str.strip().str.upper().replace({"NAN": "", "NONE": ""})
        
        # IMPORTANTE: Filtrar SOLO registros donde tipo es explícitamente S o T (sin ffill)
        work = work[work["tipo_raw"].isin(["S", "T"])].copy()
        work["tipo"] = work["tipo_raw"]
        
        # Normalizar comuna2 y nombre
        work["comuna2"] = pd.to_numeric(work["comuna2"], errors="coerce")
        work["nombre_ccgrd"] = work["nombre_ccgrd"].astype(str).str.strip()
        
        # Filtros finales
        work = work[(work["nombre_ccgrd"] != "") & (~work["nombre_ccgrd"].str.lower().isin(["nan", "none"]))]
        work = work[work["comuna2"].notna()]
        
        # Iterar sobre registros únicos
        for _, row in work.iterrows():
            nombre = _normalize_comite_name(row["nombre_ccgrd"])
            if not nombre:
                continue

            comuna_cod = int(row["comuna2"])
            if comuna_cod <= 0:
                continue

            rows.append({
                "comuna_cod": comuna_cod,
                "comite_comision_nombre": nombre,
            })

    if not rows:
        return None

    work = pd.DataFrame(rows)
    # NO deduplicar: mantener TODOS los registros del consolidado, incluyendo duplicados intra-hoja
    # (ej: Los Mangos y Villatina en C8 que aparecen 2 veces)
    # Esto asegura que el CSV tenga 158 filas, matching con RESUMEN COUNTIF
    # work = work.drop_duplicates(subset=["comuna_cod", "comite_comision_nombre"]).copy()
    work["comite_comision_etiqueta"] = work.apply(
        lambda r: f"C{int(r['comuna_cod'])} - {r['comite_comision_nombre']}", axis=1
    )
    work = work.sort_values(["comuna_cod", "comite_comision_nombre"]).reset_index(drop=True)
    return work[expected_cols]


def _build_dim_comites_desde_excel_maestro() -> pd.DataFrame | None:
    expected_cols = ["comuna_cod", "comite_comision_nombre", "comite_comision_etiqueta"]
    if not EXCEL_MAESTRO_POWERBI.exists():
        return None

    try:
        raw = pd.read_excel(EXCEL_MAESTRO_POWERBI, sheet_name="Comites")
    except Exception:
        return None

    col_cod = "Comuna_Cod" if "Comuna_Cod" in raw.columns else None
    col_comuna = "Comuna" if "Comuna" in raw.columns else None
    col_nombre = "Nombre_Comite" if "Nombre_Comite" in raw.columns else None
    col_estado = "Estado" if "Estado" in raw.columns else None

    if col_nombre is None:
        return None

    work = pd.DataFrame()
    work["comuna_cod"] = raw.apply(
        lambda row: _parse_comuna_code(
            row.get(col_cod) if col_cod else None,
            row.get(col_comuna) if col_comuna else None,
        ),
        axis=1,
    )
    work["comite_comision_nombre"] = raw[col_nombre].apply(_normalize_comite_name)

    if col_estado:
        estado_norm = raw[col_estado].astype(str).str.strip().str.lower()
        work = work[~estado_norm.isin({"inactivo", "inactive"})]

    work = work[work["comuna_cod"].notna()]
    work["comuna_cod"] = work["comuna_cod"].astype(int)
    work = work[(work["comuna_cod"] > 0) & (work["comite_comision_nombre"] != "")]
    work = work.drop_duplicates(subset=["comuna_cod", "comite_comision_nombre"]).copy()
    work["comite_comision_etiqueta"] = work.apply(
        lambda r: f"C{int(r['comuna_cod'])} - {r['comite_comision_nombre']}", axis=1
    )
    work = work.sort_values(["comuna_cod", "comite_comision_nombre"]).reset_index(drop=True)
    return work[expected_cols]


def build_satc_tables() -> tuple[pd.DataFrame, pd.DataFrame]:
    if SATC_VALID.exists():
        satc_raw = pd.read_excel(SATC_VALID, sheet_name="SATC_Validacion")
        satc_df = pd.DataFrame(
            {
                "SATC_ID": satc_raw["satc_codigo"].astype(str),
                "SATC_Nombre": satc_raw["satc_nombre"].astype(str),
                "Comuna_Cod": pd.to_numeric(satc_raw["comuna_cod"], errors="coerce").fillna(0).astype(int),
                "Talleres": pd.to_numeric(satc_raw["talleres_registros"], errors="coerce").fillna(0).astype(int),
                "Comites": pd.to_numeric(satc_raw["comites_registros"], errors="coerce").fillna(0).astype(int),
                "Activo": satc_raw["existe_datos"].map({True: "Sí", False: "No"}).fillna("Sí"),
            }
        )
    else:
        satc_df = pd.DataFrame(columns=["SATC_ID", "SATC_Nombre", "Comuna_Cod", "Talleres", "Comites", "Activo"])

    rel_df = (
        satc_df.groupby("Comuna_Cod", dropna=False)
        .size()
        .reset_index(name="SATC_Cantidad")
        .sort_values("Comuna_Cod")
    )
    return satc_df, rel_df


def build_semilleros_confiable(dim_sem_df: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame()
    sem_id = pd.to_numeric(dim_sem_df.get("N°"), errors="coerce")
    fallback = pd.Series(range(1, len(dim_sem_df) + 1), index=dim_sem_df.index, dtype="float64")
    out["semillero_id"] = sem_id.fillna(fallback).astype(int)
    out["semillero_nombre"] = dim_sem_df.get("Semillero", "").astype(str)
    out["comuna_cod"] = pd.to_numeric(dim_sem_df.get("Comuna"), errors="coerce").fillna(0).astype(int)
    out["comuna_nombre"] = dim_sem_df.get("Comuna_Nombre", "").astype(str)
    out["barrio_organizacion"] = dim_sem_df.get("Barrio_Organización", "").astype(str)
    out["fuente_semillero"] = "DAGRD"
    out["activo"] = "Sí"
    return out


def build_dim_semilleros_vacia() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "N°",
            "Semillero",
            "Comuna",
            "Comuna_Nombre",
            "Barrio_Organización",
            "Total_Semilleros",
        ]
    )


def build_dim_comites_desde_fuente() -> pd.DataFrame:
    # Prioridad 1: fuente confiable solicitada por usuario (consolidado por comuna/corregimiento).
    from_consolidado = _build_dim_comites_desde_consolidado_confiable()
    if from_consolidado is not None and len(from_consolidado) > 0:
        from_consolidado.to_csv(DIM_COMITES_CSV, index=False, encoding="utf-8-sig")
        print(f"Fuente comites usada: {_get_consolidado_comites_path()}")
        print(f"CSV actualizado: {DIM_COMITES_CSV} ({len(from_consolidado)} filas)")
        return from_consolidado

    # Prioridad 2: base consolidada para Power BI (usa Comuna/Comuna_Cod).
    from_maestro = _build_dim_comites_desde_excel_maestro()
    if from_maestro is not None and len(from_maestro) > 0:
        from_maestro.to_csv(DIM_COMITES_CSV, index=False, encoding="utf-8-sig")
        print(f"Fuente comites usada: {EXCEL_MAESTRO_POWERBI} [Comites]")
        print(f"CSV actualizado: {DIM_COMITES_CSV} ({len(from_maestro)} filas)")
        return from_maestro

    patterns = ["*comites*comisiones*2026*.xls*", "*comit*comision*2026*.xls*"]
    candidates: list[Path] = []
    for pattern in patterns:
        candidates.extend((BASE / "data" / "model").glob(pattern))
    candidates = sorted({p for p in candidates if p.is_file()}, key=lambda p: p.stat().st_mtime, reverse=True)

    expected_cols = ["comuna_cod", "comite_comision_nombre", "comite_comision_etiqueta"]

    if not candidates:
        if DIM_COMITES_CSV.exists():
            return pd.read_csv(DIM_COMITES_CSV)
        return pd.DataFrame(columns=expected_cols)

    source = candidates[0]
    try:
        raw = pd.read_excel(source, sheet_name=0)
    except Exception:
        if DIM_COMITES_CSV.exists():
            return pd.read_csv(DIM_COMITES_CSV)
        return pd.DataFrame(columns=expected_cols)

    comuna_col = "comuna_cod" if "comuna_cod" in raw.columns else None
    comuna_name_col = "comuna_nom" if "comuna_nom" in raw.columns else None
    nombre_col = "nom_titulo" if "nom_titulo" in raw.columns else None
    estado_col = "estado" if "estado" in raw.columns else None

    if nombre_col is None:
        if DIM_COMITES_CSV.exists():
            return pd.read_csv(DIM_COMITES_CSV)
        return pd.DataFrame(columns=expected_cols)

    work = pd.DataFrame()
    work["comuna_cod"] = raw.apply(
        lambda row: _parse_comuna_code(
            row.get(comuna_col) if comuna_col else None,
            row.get(comuna_name_col) if comuna_name_col else None,
        ),
        axis=1,
    )
    work["comite_comision_nombre"] = raw[nombre_col].apply(_normalize_comite_name)

    if estado_col:
        estado_norm = raw[estado_col].astype(str).str.strip().str.lower()
        work = work[~estado_norm.isin({"inactivo", "inactive"})]

    work = work[work["comuna_cod"].notna()]
    work["comuna_cod"] = work["comuna_cod"].astype(int)
    work = work[(work["comuna_cod"] > 0) & (work["comite_comision_nombre"] != "")]
    work = work.drop_duplicates(subset=["comuna_cod", "comite_comision_nombre"]).copy()
    work["comite_comision_etiqueta"] = work.apply(
        lambda r: f"C{int(r['comuna_cod'])} - {r['comite_comision_nombre']}", axis=1
    )
    work = work.sort_values(["comuna_cod", "comite_comision_nombre"]).reset_index(drop=True)

    work.to_csv(DIM_COMITES_CSV, index=False, encoding="utf-8-sig")
    print(f"Fuente comites usada: {source}")
    print(f"CSV actualizado: {DIM_COMITES_CSV} ({len(work)} filas)")
    return work[expected_cols]


def write_df_to_sheet(wb, sheet_name: str, df: pd.DataFrame) -> None:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)

    ws = wb.create_sheet(title=sheet_name[:31])
    ws.append(list(df.columns))

    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))


def read_sheet_df(path: Path, sheet_name: str, fallback: pd.DataFrame) -> pd.DataFrame:
    try:
        return pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
    except Exception:
        return fallback.copy()


def main() -> None:
    if not MODELO.exists():
        raise FileNotFoundError(f"No existe archivo modelo: {MODELO}")

    wb = load_workbook(MODELO)
    existing_sheets = set(wb.sheetnames)

    satc_existing = read_sheet_df(MODELO, "Dim_SATC", pd.DataFrame())
    sem_existing = read_sheet_df(MODELO, "Dim_Semilleros", pd.DataFrame())

    # SATC requeridos por Power BI
    satc_df, rel_df = build_satc_tables()

    # Blindaje: nunca degradar el catálogo SATC por debajo de 37
    if len(satc_df) < 37:
        if not satc_existing.empty and len(satc_existing) >= 37:
            satc_df = satc_existing.copy()
        elif ENTREGABLE_MODELO.exists():
            try:
                satc_fallback = pd.read_excel(ENTREGABLE_MODELO, sheet_name="Dim_SATC", engine="openpyxl")
                if len(satc_fallback) >= 37:
                    satc_df = satc_fallback
            except Exception:
                pass

        rel_df = (
            satc_df.groupby("Comuna_Cod", dropna=False)
            .size()
            .reset_index(name="SATC_Cantidad")
            .sort_values("Comuna_Cod")
        )

    write_df_to_sheet(wb, "Dim_SATC", satc_df)
    write_df_to_sheet(wb, "Dim_SATC_Relaciones", rel_df)

    # Semilleros requeridos por Power BI
    if "Dim_Semilleros" not in existing_sheets:
        dim_semilleros_df = build_dim_semilleros_vacia()
        if ENTREGABLE_MODELO.exists():
            try:
                dim_semilleros_df = pd.read_excel(ENTREGABLE_MODELO, sheet_name="Dim_Semilleros", engine="openpyxl")
            except Exception:
                dim_semilleros_df = build_dim_semilleros_vacia()
        write_df_to_sheet(wb, "Dim_Semilleros", dim_semilleros_df)
    else:
        dim_semilleros_df = sem_existing if not sem_existing.empty else build_dim_semilleros_vacia()

    sem_conf = build_semilleros_confiable(dim_semilleros_df)
    write_df_to_sheet(wb, "Dim_Semilleros_Confiable", sem_conf)

    # Comités/Comisiones requeridos por consultas del PBIX
    write_df_to_sheet(wb, "Dim_Comites_Comisiones_2026", build_dim_comites_desde_fuente())

    wb.save(MODELO)

    print(f"OK: {MODELO}")
    print("Hojas críticas garantizadas:")
    final_sheets = set(wb.sheetnames)
    for n in ["Dim_SATC", "Dim_SATC_Relaciones", "Dim_Semilleros", "Dim_Semilleros_Confiable", "Dim_Comites_Comisiones_2026", "Hecho_Participacion_General", "Hecho_Demografia", "Hecho_Simulacros"]:
        print(f" - {n}: {'SI' if n in final_sheets else 'NO'}")


if __name__ == "__main__":
    main()

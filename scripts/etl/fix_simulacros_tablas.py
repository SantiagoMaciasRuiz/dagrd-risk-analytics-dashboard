"""
Asegura que el libro fuente contenga ambas hojas: 'Tablas dinámicas' y 'Simulacros'.
Si solo existe 'Tablas dinámicas', crea una copia como 'Simulacros'.
"""
from pathlib import Path
from openpyxl import load_workbook

p = Path(r"data/source/Reporte de actividades equipo social 2026.xlsx")
if not p.exists():
    print("Archivo fuente no encontrado:", p)
    raise SystemExit(2)

wb = load_workbook(p)
names = wb.sheetnames
print('Hojas actuales:', names)
if 'Tablas dinámicas' in names and 'Simulacros' not in names:
    ws = wb['Tablas dinámicas']
    new = wb.copy_worksheet(ws)
    new.title = 'Simulacros'
    wb.save(p)
    print("Copia creada: 'Simulacros' desde 'Tablas dinámicas'.")
else:
    print('No se requiere acción.')

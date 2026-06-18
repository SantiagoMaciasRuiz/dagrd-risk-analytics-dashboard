#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path
from typing import Dict, Iterable, List

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


BASE_DIR = Path(r"c:\Users\santi\OneDrive\Escritorio\Chamba\Dashboard")
DATA_DIR = BASE_DIR / "datos"
OUTPUT_FILE = BASE_DIR / "Modelo_DAGRD_PowerBI_Limpio.xlsx"
EXCEL_MIN_DATE = pd.Timestamp("1900-01-01")


SOURCE_FILES = {
    "talleres": "talleres_2026-03-10_19-14.xls",
    "comites": "comités_comisiones comunitarios_2026-03-10_19-15.xls",
    "eventos": "eventos históricos_2026-03-10_19-15.xls",
    "cosegrd": "cosegrd_2026-03-10_19-15.xls",
    "estudios": "estudios_2026-03-10_19-15.xls",
    "instituciones": "instituciones educativas (puntos)_2026-03-10_19-15.xls",
    "obras": "obras_2026-03-10_19-14.xls",
}


MONTHS_ES = {
    1: "enero",
    2: "febrero",
    3: "marzo",
    4: "abril",
    5: "mayo",
    6: "junio",
    7: "julio",
    8: "agosto",
    9: "septiembre",
    10: "octubre",
    11: "noviembre",
    12: "diciembre",
}


DAY_NAMES_ES = {
    0: "lunes",
    1: "martes",
    2: "miercoles",
    3: "jueves",
    4: "viernes",
    5: "sabado",
    6: "domingo",
}


def read_xls(file_key: str) -> pd.DataFrame:
    path = DATA_DIR / SOURCE_FILES[file_key]
    return pd.read_excel(path)


def trim_text_columns(df: pd.DataFrame) -> pd.DataFrame:
    text_cols = df.select_dtypes(include=["object"]).columns
    for col in text_cols:
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].replace({"nan": np.nan, "None": np.nan})
    return df


def parse_date_series(series: pd.Series) -> pd.Series:
    if pd.api.types.is_numeric_dtype(series):
        year_values = pd.to_numeric(series, errors="coerce")
        year_mask = year_values.between(1900, 2100)
        out = pd.Series(pd.NaT, index=series.index, dtype="datetime64[ns]")
        out.loc[year_mask] = pd.to_datetime(
            year_values.loc[year_mask].astype(int).astype(str) + "-01-01",
            errors="coerce",
        )
        return out

    parsed = pd.to_datetime(series, errors="coerce")

    numeric_values = pd.to_numeric(series, errors="coerce")
    year_mask = numeric_values.between(1900, 2100)
    if year_mask.any():
        parsed.loc[year_mask] = pd.to_datetime(
            numeric_values.loc[year_mask].astype(int).astype(str) + "-01-01",
            errors="coerce",
        )
    # Excel muestra #### cuando una fecha es anterior a 1900.
    parsed = parsed.where(parsed.isna() | (parsed >= EXCEL_MIN_DATE), pd.NaT)
    return parsed


def add_time_columns(df: pd.DataFrame, date_col: str = "fecha") -> pd.DataFrame:
    df[date_col] = parse_date_series(df[date_col])
    df["anio"] = df[date_col].dt.year
    df["mes_num"] = df[date_col].dt.month
    df["mes_nombre"] = df["mes_num"].map(MONTHS_ES)
    return df


def to_numeric(df: pd.DataFrame, cols: Iterable[str]) -> pd.DataFrame:
    for col in cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def keep_columns(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    safe_cols = [c for c in cols if c in df.columns]
    return df[safe_cols].copy()


def build_fact_talleres() -> pd.DataFrame:
    raw = trim_text_columns(read_xls("talleres"))
    cols = [
        "id",
        "fecha_ini",
        "comuna_cod",
        "comuna_nom",
        "barrio_cod",
        "barrio_nom",
        "latitud",
        "longitud",
        "instancia",
        "tipo_activ",
        "modalidad",
        "partic_num",
        "mujere_num",
        "hombre_num",
        "ninos_num",
        "impac_indi",
        "pri_infanc",
        "nino_adole",
        "juventud",
        "adulto",
        "adulto_may",
        "discapacid",
        "afrodescen",
        "campesino",
        "pob_victim",
        "pob_migran",
        "pob_lgtbi",
        "pob_indige",
        "pob_rom",
        "universidades",
        "colegios",
        "entida_aso",
        "cert_activ",
        "rel_metPDD",
        "profesionales",
    ]
    df = keep_columns(raw, cols)
    df = df.rename(
        columns={
            "id": "id_taller",
            "fecha_ini": "fecha",
            "comuna_nom": "comuna",
            "barrio_nom": "barrio",
            "tipo_activ": "tipo_actividad",
            "partic_num": "participantes",
            "mujere_num": "mujeres",
            "hombre_num": "hombres",
            "ninos_num": "ninos",
            "impac_indi": "impacto_indirecto",
            "entida_aso": "entidad_asociada",
            "cert_activ": "cert_actividad",
            "rel_metPDD": "rel_metpdd",
        }
    )

    df = add_time_columns(df, "fecha")

    numeric_cols = [
        "id_taller",
        "comuna_cod",
        "barrio_cod",
        "latitud",
        "longitud",
        "participantes",
        "mujeres",
        "hombres",
        "ninos",
        "impacto_indirecto",
        "pri_infanc",
        "nino_adole",
        "juventud",
        "adulto",
        "adulto_may",
        "discapacid",
        "afrodescen",
        "campesino",
        "pob_victim",
        "pob_migran",
        "pob_lgtbi",
        "pob_indige",
        "pob_rom",
        "universidades",
        "colegios",
    ]
    df = to_numeric(df, numeric_cols)
    return df


def build_fact_comites() -> pd.DataFrame:
    raw = trim_text_columns(read_xls("comites"))
    cols = [
        "id",
        "comuna_cod",
        "comuna_nom",
        "barrio_cod",
        "barrio_nom",
        "latitud",
        "longitud",
        "nom_titulo",
        "junt_accio",
        "direccion",
        "coordi_nom",
        "coo_contac",
        "estado",
        "integr_num",
    ]
    df = keep_columns(raw, cols)
    df = df.rename(
        columns={
            "id": "id_comite",
            "comuna_nom": "comuna",
            "barrio_nom": "barrio",
            "nom_titulo": "nombre_comite",
            "junt_accio": "junta_accion",
            "coordi_nom": "coordinador",
            "coo_contac": "contacto",
            "integr_num": "integrantes",
        }
    )
    numeric_cols = [
        "id_comite",
        "comuna_cod",
        "barrio_cod",
        "latitud",
        "longitud",
        "integrantes",
    ]
    df = to_numeric(df, numeric_cols)
    return df


def build_fact_eventos() -> pd.DataFrame:
    raw = trim_text_columns(read_xls("eventos"))
    cols = [
        "id",
        "comuna_cod",
        "comuna_nom",
        "barrio_cod",
        "barrio_nom",
        "latitud",
        "longitud",
        "nom_titulo",
        "fecha",
        "localizaci",
        "nivel",
        "descripcio",
        "enlaces",
        "vict_fatal",
        "lesionados",
        "fenomenos",
    ]
    df = keep_columns(raw, cols)
    df = df.rename(
        columns={
            "id": "id_evento",
            "comuna_nom": "comuna",
            "barrio_nom": "barrio",
            "nom_titulo": "nombre_evento",
            "localizaci": "localizacion",
            "descripcio": "descripcion",
            "vict_fatal": "victimas_fatales",
            "fenomenos": "fenomeno",
        }
    )
    df["fecha_fuente"] = df["fecha"]
    df["anio_fuente"] = pd.to_datetime(df["fecha_fuente"], errors="coerce").dt.year
    df = add_time_columns(df, "fecha")
    numeric_cols = [
        "id_evento",
        "comuna_cod",
        "barrio_cod",
        "latitud",
        "longitud",
        "victimas_fatales",
        "lesionados",
        "anio_fuente",
    ]
    df = to_numeric(df, numeric_cols)
    return df


def build_fact_cosegrd() -> pd.DataFrame:
    raw = trim_text_columns(read_xls("cosegrd"))
    cols = [
        "id",
        "comuna_cod",
        "comuna_nom",
        "barrio_cod",
        "barrio_nom",
        "latitud",
        "longitud",
        "nom_titulo",
        "integrante",
        "estado",
    ]
    df = keep_columns(raw, cols)
    df = df.rename(
        columns={
            "id": "id_cosegrd",
            "comuna_nom": "comuna",
            "barrio_nom": "barrio",
            "nom_titulo": "nombre_esquema",
        }
    )
    numeric_cols = [
        "id_cosegrd",
        "comuna_cod",
        "barrio_cod",
        "latitud",
        "longitud",
    ]
    df = to_numeric(df, numeric_cols)
    return df


def build_fact_estudios() -> pd.DataFrame:
    raw = trim_text_columns(read_xls("estudios"))
    cols = [
        "id",
        "comuna_cod",
        "comuna_nom",
        "barrio_cod",
        "barrio_nom",
        "latitud",
        "longitud",
        "nom_titulo",
        "contra_num",
        "objeto",
        "tipo_estud",
        "estado",
        "descripcion",
        "fecha",
        "contratist",
        "intervento",
        "ent_ejecut",
        "fenomenos",
    ]
    df = keep_columns(raw, cols)
    df = df.rename(
        columns={
            "id": "id_estudio",
            "comuna_nom": "comuna",
            "barrio_nom": "barrio",
            "nom_titulo": "nombre_estudio",
            "contra_num": "contrato_num",
            "tipo_estud": "tipo_estudio",
            "contratist": "contratista",
            "intervento": "interventor",
            "ent_ejecut": "entidad_ejecutora",
            "fenomenos": "fenomeno",
        }
    )
    df = add_time_columns(df, "fecha")
    numeric_cols = [
        "id_estudio",
        "comuna_cod",
        "barrio_cod",
        "latitud",
        "longitud",
    ]
    df = to_numeric(df, numeric_cols)
    return df


def build_fact_instituciones() -> pd.DataFrame:
    raw = trim_text_columns(read_xls("instituciones"))
    cols = [
        "id",
        "estado",
        "nucleo",
        "sector",
        "latitud",
        "longitud",
        "servicio",
        "vigencia",
        "direccion",
        "codigo_dane",
        "clasificacion",
        "codigo_barrio",
        "codigo_comuna",
        "nombre_barrio",
        "nombre_comuna",
        "sede_educativa",
        "nombre_establecimiento",
        "institucion",
    ]
    df = keep_columns(raw, cols)
    df = df.rename(
        columns={
            "id": "id_ie",
            "codigo_barrio": "barrio_cod",
            "codigo_comuna": "comuna_cod",
            "nombre_barrio": "barrio",
            "nombre_comuna": "comuna",
        }
    )
    numeric_cols = [
        "id_ie",
        "latitud",
        "longitud",
        "vigencia",
        "barrio_cod",
        "comuna_cod",
    ]
    df = to_numeric(df, numeric_cols)
    return df


def build_fact_obras() -> pd.DataFrame:
    raw = trim_text_columns(read_xls("obras"))
    cols = [
        "id",
        "comuna_cod",
        "comuna_nom",
        "barrio_cod",
        "barrio_nom",
        "direccion",
        "latitud",
        "longitud",
        "nom_titulo",
        "contra_num",
        "objeto",
        "tipo_obra",
        "estado",
        "avance",
        "descripcion",
        "fecha",
        "contratist",
        "intervento",
        "ent_ejecut",
        "fenomenos",
    ]
    df = keep_columns(raw, cols)
    df = df.rename(
        columns={
            "id": "id_obra",
            "comuna_nom": "comuna",
            "barrio_nom": "barrio",
            "nom_titulo": "nombre_obra",
            "contra_num": "contrato_num",
            "contratist": "contratista",
            "intervento": "interventor",
            "ent_ejecut": "entidad_ejecutora",
            "fenomenos": "fenomeno",
        }
    )
    df = add_time_columns(df, "fecha")
    numeric_cols = [
        "id_obra",
        "comuna_cod",
        "barrio_cod",
        "latitud",
        "longitud",
        "avance",
    ]
    df = to_numeric(df, numeric_cols)
    return df


def build_dim_comuna(facts: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    pairs = []
    for df in facts.values():
        if {"comuna_cod", "comuna"}.issubset(df.columns):
            pairs.append(df[["comuna_cod", "comuna"]])

    dim = pd.concat(pairs, ignore_index=True)
    dim = dim.dropna(subset=["comuna_cod", "comuna"])
    dim = dim.drop_duplicates().copy()
    dim["comuna_cod"] = pd.to_numeric(dim["comuna_cod"], errors="coerce")
    dim = dim.dropna(subset=["comuna_cod"])
    dim["comuna_cod"] = dim["comuna_cod"].astype(int)
    dim["tipo_territorio"] = np.where(
        dim["comuna_cod"].between(1, 16), "Comuna", "Corregimiento"
    )
    dim = dim.sort_values(["comuna_cod", "comuna"]).reset_index(drop=True)
    return dim


def map_instancia_to_seccion(instancia: str) -> str:
    value = str(instancia).lower()
    if "comunit" in value:
        return "Comunitaria"
    if "empres" in value:
        return "Empresarial"
    if "instit" in value or "gestion interna" in value:
        return "Institucional"
    if "educ" in value or "primera infancia" in value or "buen comienzo" in value:
        return "Educativa"
    return "Otros"


def build_dim_instancia(fact_talleres: pd.DataFrame) -> pd.DataFrame:
    values = sorted([x for x in fact_talleres["instancia"].dropna().unique()])
    rows = []
    for idx, value in enumerate(values, start=1):
        rows.append(
            {
                "id_instancia": idx,
                "instancia": value,
                "seccion_tablero": map_instancia_to_seccion(value),
            }
        )
    return pd.DataFrame(rows)


def map_grupo_actividad(tipo: str) -> str:
    value = str(tipo).lower()
    if "simulacro" in value:
        return "Simulacros"
    if "capacit" in value:
        return "Capacitacion"
    if "taller" in value:
        return "Taller"
    if "socializacion" in value:
        return "Socializacion"
    if "acompan" in value:
        return "Acompanamiento"
    return "Otros"


def build_dim_actividad(fact_talleres: pd.DataFrame) -> pd.DataFrame:
    values = sorted([x for x in fact_talleres["tipo_actividad"].dropna().unique()])
    rows = []
    for idx, value in enumerate(values, start=1):
        rows.append(
            {
                "id_actividad": idx,
                "tipo_actividad": value,
                "grupo_actividad": map_grupo_actividad(value),
            }
        )
    return pd.DataFrame(rows)


def build_dim_poblacion() -> pd.DataFrame:
    rows = [
        (1, "pri_infanc", "Primera infancia", "Edad"),
        (2, "nino_adole", "Nino adolescencia", "Edad"),
        (3, "juventud", "Juventud", "Edad"),
        (4, "adulto", "Adulto", "Edad"),
        (5, "adulto_may", "Adulto mayor", "Edad"),
        (6, "discapacid", "Discapacidad", "Condicion"),
        (7, "afrodescen", "Afrodescendiente", "Etnia"),
        (8, "campesino", "Campesino", "Territorio"),
        (9, "pob_victim", "Poblacion victima", "Condicion"),
        (10, "pob_migran", "Poblacion migrante", "Condicion"),
        (11, "pob_lgtbi", "Poblacion lgtbi", "Identidad"),
        (12, "pob_indige", "Poblacion indigena", "Etnia"),
        (13, "pob_rom", "Poblacion rom", "Etnia"),
    ]
    return pd.DataFrame(
        rows,
        columns=["id_poblacion", "codigo_campo", "poblacion", "segmento"],
    )


def build_dim_fecha(date_series: List[pd.Series]) -> pd.DataFrame:
    joined = pd.concat(date_series, ignore_index=True)
    joined = pd.to_datetime(joined, errors="coerce").dropna()
    joined = joined[joined >= EXCEL_MIN_DATE]
    if joined.empty:
        min_date = EXCEL_MIN_DATE
        max_date = pd.Timestamp("2026-12-31")
    else:
        min_date = joined.min()
        max_date = joined.max()

    all_dates = pd.date_range(min_date, max_date, freq="D")
    dim = pd.DataFrame({"fecha": all_dates})
    dim["anio"] = dim["fecha"].dt.year
    dim["trimestre"] = "T" + dim["fecha"].dt.quarter.astype(str)
    dim["mes_num"] = dim["fecha"].dt.month
    dim["mes_nombre"] = dim["mes_num"].map(MONTHS_ES)
    dim["semana_iso"] = dim["fecha"].dt.isocalendar().week.astype(int)
    dim["dia"] = dim["fecha"].dt.day
    dim["dia_semana_num"] = dim["fecha"].dt.dayofweek
    dim["dia_semana"] = dim["dia_semana_num"].map(DAY_NAMES_ES)
    dim["es_fin_semana"] = dim["dia_semana_num"].isin([5, 6]).astype(int)
    return dim


def build_meta_base() -> pd.DataFrame:
    rows = [
        (1, "Total", "participantes", 255),
        (2, "Comunitaria", "participantes", 255),
        (3, "Educativa", "participantes", 255),
        (4, "Empresarial", "participantes", 255),
        (5, "Institucional", "participantes", 255),
    ]
    return pd.DataFrame(
        rows,
        columns=["id_meta", "seccion_tablero", "indicador", "meta_valor"],
    )


def build_resumen_carga(dataframes: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    for name, df in dataframes.items():
        rows.append(
            {
                "tabla": name,
                "filas": int(df.shape[0]),
                "columnas": int(df.shape[1]),
                "nulos_totales": int(df.isna().sum().sum()),
            }
        )
    return pd.DataFrame(rows).sort_values("tabla").reset_index(drop=True)


def safe_table_name(sheet_name: str) -> str:
    cleaned = "".join(ch if ch.isalnum() else "_" for ch in sheet_name)
    cleaned = cleaned.strip("_")
    if not cleaned:
        cleaned = "Tabla"
    return f"tbl_{cleaned}"[:255]


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        if ws.max_row < 1 or ws.max_column < 1:
            continue

        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 300), min_col=col_idx, max_col=col_idx):
                value = row[0].value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))
            width = min(max(max_len + 2, 10), 45)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        if ws.max_row >= 2:
            table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            table = Table(displayName=safe_table_name(ws.title), ref=table_ref)
            table_style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = table_style
            ws.add_table(table)

    wb.save(path)


def main() -> None:
    fact_talleres = build_fact_talleres()
    fact_comites = build_fact_comites()
    fact_eventos = build_fact_eventos()
    fact_cosegrd = build_fact_cosegrd()
    fact_estudios = build_fact_estudios()
    fact_instituciones = build_fact_instituciones()
    fact_obras = build_fact_obras()

    facts = {
        "Fact_Talleres": fact_talleres,
        "Fact_Comites": fact_comites,
        "Fact_Eventos": fact_eventos,
        "Fact_Cosegrd": fact_cosegrd,
        "Fact_Estudios": fact_estudios,
        "Fact_Instituciones": fact_instituciones,
        "Fact_Obras": fact_obras,
    }

    dim_comuna = build_dim_comuna(facts)
    dim_instancia = build_dim_instancia(fact_talleres)
    dim_actividad = build_dim_actividad(fact_talleres)
    dim_poblacion = build_dim_poblacion()
    dim_fecha = build_dim_fecha(
        [
            fact_talleres["fecha"],
            fact_eventos["fecha"],
            fact_estudios["fecha"],
            fact_obras["fecha"],
        ]
    )
    meta_base = build_meta_base()

    output_tables: Dict[str, pd.DataFrame] = {
        "Fact_Talleres": fact_talleres,
        "Fact_Comites": fact_comites,
        "Fact_Eventos": fact_eventos,
        "Fact_Cosegrd": fact_cosegrd,
        "Fact_Estudios": fact_estudios,
        "Fact_Instituciones": fact_instituciones,
        "Fact_Obras": fact_obras,
        "Dim_Fecha": dim_fecha,
        "Dim_Comuna": dim_comuna,
        "Dim_Instancia": dim_instancia,
        "Dim_Actividad": dim_actividad,
        "Dim_Poblacion": dim_poblacion,
        "Meta_Base": meta_base,
    }

    resumen = build_resumen_carga(output_tables)
    output_tables = {"Resumen_Carga": resumen, **output_tables}

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
        date_format="yyyy-mm-dd",
        datetime_format="yyyy-mm-dd",
    ) as writer:
        for sheet_name, df in output_tables.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    format_workbook(OUTPUT_FILE)

    print(f"OK: archivo generado en {OUTPUT_FILE}")
    for name, df in output_tables.items():
        print(f"- {name}: {df.shape[0]} filas x {df.shape[1]} columnas")


if __name__ == "__main__":
    main()